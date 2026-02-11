import logging
from aiogram import Bot, Dispatcher, types
from aiogram.types import ReplyKeyboardMarkup
from aiogram.utils import executor
from openpyxl import load_workbook
from parser import parse_inventory, find_similar
import os

TOKEN = "ВСТАВЬ_СЮДА_ТОКЕН"

logging.basicConfig(level=logging.INFO)

bot = Bot(token=TOKEN)
dp = Dispatcher(bot)

user_state = {}
user_miss = {}

# ================= КНОПКИ =================

keyboard = ReplyKeyboardMarkup(resize_keyboard=True)
keyboard.add("Загрузить таблицу программы")
keyboard.add("Получить отчет")
keyboard.add("Показать несовпадения")

# ================= START =================

@dp.message_handler(commands=['start'])
async def start(message: types.Message):
    await message.answer(
        "Привет. Я бот учета.\nВыбери действие:",
        reply_markup=keyboard
    )

# ================= КНОПКА ЗАГРУЗКИ =================

@dp.message_handler(lambda m: m.text == "Загрузить таблицу программы")
async def request_program(message: types.Message):
    user_state[message.from_user.id] = "waiting_program"
    await message.answer("Отправьте Excel файл (.xlsx) из программы")

# ================= ПРИЁМ ФАЙЛА =================

@dp.message_handler(content_types=['document'])
async def handle_excel(message: types.Message):

    if user_state.get(message.from_user.id) != "waiting_program":
        return

    document = message.document

    # ❗ ВТОРАЯ ВОЗМОЖНАЯ ПРИЧИНА — ФАЙЛ НЕ ТОГО ФОРМАТА
    if not document.file_name.endswith(".xlsx"):
        await message.answer("Нужен файл формата .xlsx")
        return

    file = await bot.get_file(document.file_id)
    await bot.download_file(file.file_path, "program.xlsx")

    # ❗ ЕСЛИ НЕТ tabl0.xlsx — создаём ошибку понятную
    if not os.path.exists("tabl0.xlsx"):
        await message.answer("Ошибка: tabl0.xlsx отсутствует в проекте")
        return

    base_wb = load_workbook("tabl0.xlsx")
    base_ws = base_wb.active

    prog_wb = load_workbook("program.xlsx")
    prog_ws = prog_wb.active

    program_data = {}

    # ❗ ЧИСТИМ НАЗВАНИЯ (ОБЯЗАТЕЛЬНО)
    for row in prog_ws.iter_rows(min_row=1, values_only=True):
        if row[0] and row[1]:
            name = str(row[0]).strip().lower()
            qty = row[1]
            program_data[name] = qty

    for row in base_ws.iter_rows(min_row=2):
        product_name = str(row[0].value).strip().lower()

        if product_name in program_data:
            row[1].value = program_data[product_name]

    base_wb.save("work.xlsx")

    user_state[message.from_user.id] = "waiting_fact"

    await message.answer("Файл принят.\nТеперь отправьте фактические остатки списком.")

# ================= ПРИЁМ ФАКТА =================

@dp.message_handler(lambda m: user_state.get(m.from_user.id) == "waiting_fact")
async def handle_fact(message: types.Message):

    data = parse_inventory(message.text)

    wb = load_workbook("work.xlsx")
    ws = wb.active

    table_products = [
        str(row[0].value).strip().lower()
        for row in ws.iter_rows(min_row=2)
    ]

    misses = []

    for name, qty in data.items():

        found_row = None

        for row in ws.iter_rows(min_row=2):
            product_name = str(row[0].value).strip().lower()

            if name == product_name:
                found_row = row
                break

        if not found_row:
            similar = find_similar(name, table_products)
            if similar:
                for row in ws.iter_rows(min_row=2):
                    if str(row[0].value).strip().lower() == similar:
                        found_row = row
                        break

        if found_row:
            found_row[2].value = qty
            plan = found_row[1].value or 0
            found_row[3].value = plan - qty
        else:
            misses.append(name)

    wb.save("report.xlsx")

    user_miss[message.from_user.id] = misses
    user_state[message.from_user.id] = None

    await message.answer("Факт принят.\nНажмите 'Получить отчет'")

# ================= ОТЧЕТ =================

@dp.message_handler(lambda m: m.text == "Получить отчет")
async def send_report(message: types.Message):

    if not os.path.exists("report.xlsx"):
        await message.answer("Сначала загрузите таблицу и введите факт")
        return

    await message.answer_document(types.InputFile("report.xlsx"))

# ================= НЕСОВПАДЕНИЯ =================

@dp.message_handler(lambda m: m.text == "Показать несовпадения")
async def show_miss(message: types.Message):

    misses = user_miss.get(message.from_user.id)

    if not misses:
        await message.answer("Несовпадений нет")
        return

    text = "Не найдено:\n\n" + "\n".join(misses)
    await message.answer(text)

# ================= ЗАПУСК =================

if __name__ == "__main__":
    executor.start_polling(dp, skip_updates=True)
